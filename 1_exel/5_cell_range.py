from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11): # 10개 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

# col_B = ws["B"] #영어 column만 가지고 오기
# for cell in col_B:
#     print(cell.value)

# col_range = ws["B:C"] #영어, 수학 column 함께 가지고 오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# row_title = ws[1] #1번째 row만 가지고 오기
# row_range = ws[2:6] #2번째 줄 ~ 6번째 줄까지 가지고 오기

# row_range = ws[2:ws.max_row] #2번째 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ") #셀의 좌표 정보를 가져올 수 있음
#         xy = coordinate_from_string(cell.coordinate) #튜플 형태로 셀의 좌표를 가져올 수 있다,
#         # print(xy, end=" ") #('A', '1')
#         print(xy[0], end="") #column 좌표 (A)
#         print(xy[1], end=" ") #row 좌표 (1)
#     print()



#전체 rows
print(tuple(ws.rows)) #(A1, B1, C1), (A2, B2, C2), ...

#전체 columns
print(tuple(ws.columns)) #(A1, A2, A3, ...), (B1, B2, B3, ...), ...

for row in tuple(ws.rows):
    print(row[1].value) # B1, B2, B3, ... 의 정보를 출력

for column in tuple(ws.columns):
    print(column[0].value) #A1, B1, C1의 정보를 출력

for row in ws.iter_rows(): # 전체 row
    print(row[1].value) # B1, B2, B3, ... 의 정보를 출력

wb.save("sample.xlsx")