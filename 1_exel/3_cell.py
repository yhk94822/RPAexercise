from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) #A1 셀의 정보를 출력
print(ws["A1"].value) #A1 셀의 값을 출력

print(ws["A10"].value) #값이 없을 땐 'None'을 출력

#row = 1, 2, 3, ...
#column = A=1, B=2, C=3, ...
print(ws.cell(column=1, row=1).value) #A1 value
print(ws.cell(column=2, row=1).value) #B1 value

c1 = ws.cell(column=3, row=1, value=10) # c1 = ws["C1"]
print(c1.value) # ws["C1"].value

index=1
#반복문을 이용해서 랜덤 숫자 채우기
for column in range(1, 11):
    for row in range(1, 11):
        # ws.cell(row=row, column=column, value=randint(0, 100)) # 0~100 사이의 숫자
        ws.cell(row=row, column=column, value = index)
        index+=1

wb.save("sample.xlsx")